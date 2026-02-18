import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
import math

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1F4E79")
subheader_fill = PatternFill("solid", fgColor="2E75B6")
subheader_font = Font(bold=True, color="FFFFFF", size=10)
label_fill = PatternFill("solid", fgColor="D6E4F0")
label_font = Font(bold=True, size=10)
normal_font = Font(size=10)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")

thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell, text, fill=None, font=None):
    cell.value = text
    cell.font = font or header_font
    cell.fill = fill or header_fill
    cell.alignment = center
    cell.border = border

def style_cell(cell, value=None, bold=False, fill=None, align=None):
    if value is not None:
        cell.value = value
    cell.font = Font(bold=bold, size=10)
    if fill:
        cell.fill = fill
    cell.alignment = align or center
    cell.border = border

# ─────────────────────────────────────────────
# SHEET 1: Q1 - Graphs
# ─────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Q1 - Graphs"

# Title
ws1.merge_cells("A1:F1")
ws1["A1"].value = "Q1 - Plot of Polynomial Equations (Domain: -10 ≤ x ≤ 10)"
ws1["A1"].font = Font(bold=True, size=13, color="FFFFFF")
ws1["A1"].fill = header_fill
ws1["A1"].alignment = center

# Column headers
headers_q1 = ["x", "y1 = x³-2x²+5x+2", "y2 = 3x³-x²+5x-3", "y3 = x²-5x+3", "y4 = 2x-9"]
for col, h in enumerate(headers_q1, 1):
    cell = ws1.cell(row=2, column=col)
    style_header(cell, h, fill=subheader_fill, font=subheader_font)
    ws1.column_dimensions[get_column_letter(col)].width = 22

# Data rows
for i, x in enumerate(range(-10, 11), start=3):
    ws1.cell(row=i, column=1).value = x
    ws1.cell(row=i, column=2).value = x**3 - 2*x**2 + 5*x + 2
    ws1.cell(row=i, column=3).value = 3*x**3 - x**2 + 5*x - 3
    ws1.cell(row=i, column=4).value = x**2 - 5*x + 3
    ws1.cell(row=i, column=5).value = 2*x - 9
    for col in range(1, 6):
        c = ws1.cell(row=i, column=col)
        c.font = normal_font
        c.alignment = center
        c.border = border
        if i % 2 == 0:
            c.fill = PatternFill("solid", fgColor="EBF3FB")

# Chart
chart1 = LineChart()
chart1.title = "Polynomial Functions over [-10, 10]"
chart1.style = 10
chart1.y_axis.title = "y"
chart1.x_axis.title = "x"
chart1.height = 15
chart1.width = 28

x_vals = Reference(ws1, min_col=1, min_row=3, max_row=23)
colors = ["FF0000", "00B050", "0070C0", "FF6600"]
names = ["y1", "y2", "y3", "y4"]
for col_idx in range(2, 6):
    data = Reference(ws1, min_col=col_idx, min_row=2, max_row=23)
    series = openpyxl.chart.Series(data, x_vals, title_from_data=True)
    series.graphicalProperties.line.solidFill = colors[col_idx-2]
    series.graphicalProperties.line.width = 20000
    chart1.series.append(series)

chart1.plot_area.graphicalProps = None
ws1.add_chart(chart1, "G2")

# ─────────────────────────────────────────────
# SHEET 2: Q2 - EOS
# ─────────────────────────────────────────────
ws2 = wb.create_sheet("Q2 - EOS")

# Title
ws2.merge_cells("A1:H1")
ws2["A1"].value = "Q2 - Specific Volume of Gas Mixture (25% NH3, 75% N2/H2 at 1:3 ratio)"
ws2["A1"].font = Font(bold=True, size=13, color="FFFFFF")
ws2["A1"].fill = header_fill
ws2["A1"].alignment = center

# Given data
ws2.merge_cells("A3:H3")
ws2["A3"].value = "Given Conditions"
ws2["A3"].font = Font(bold=True, size=11, color="FFFFFF")
ws2["A3"].fill = subheader_fill
ws2["A3"].alignment = center

given = [
    ("Temperature T (K)", 550),
    ("Pressure P (atm)", 270),
    ("Gas Constant R (cm³·atm/mol·K)", 82.06),
]
for i, (label, val) in enumerate(given, start=4):
    ws2.cell(row=i, column=1).value = label
    ws2.cell(row=i, column=1).font = label_font
    ws2.cell(row=i, column=1).fill = label_fill
    ws2.cell(row=i, column=1).border = border
    ws2.cell(row=i, column=1).alignment = left
    ws2.merge_cells(f"A{i}:C{i}")
    ws2.cell(row=i, column=4).value = val
    ws2.cell(row=i, column=4).font = normal_font
    ws2.cell(row=i, column=4).border = border
    ws2.cell(row=i, column=4).alignment = center

# Component table
ws2.merge_cells("A8:H8")
ws2["A8"].value = "Component Critical Properties & Composition"
ws2["A8"].font = Font(bold=True, size=11, color="FFFFFF")
ws2["A8"].fill = subheader_fill
ws2["A8"].alignment = center

comp_headers = ["Component", "yi (mol frac)", "Tc (K)", "Pc (atm)", "ω (acentric)", "Tr = T/Tc"]
for col, h in enumerate(comp_headers, 1):
    cell = ws2.cell(row=9, column=col)
    style_header(cell, h, fill=subheader_fill, font=subheader_font)
    ws2.column_dimensions[get_column_letter(col)].width = 20

T = 550
P = 270
R = 82.06
components = [
    ("NH3",  0.25,   405.6, 111.3,  0.250),
    ("N2",   0.1875, 126.2,  33.5,  0.040),
    ("H2",   0.5625,  33.2,  12.8, -0.216),
]
comp_data = []
for i, (name, yi, Tc, Pc, omega) in enumerate(components, start=10):
    Tr = T / Tc
    row_data = [name, yi, Tc, Pc, omega, round(Tr, 4)]
    comp_data.append((yi, Tc, Pc, omega, Tr))
    for col, val in enumerate(row_data, 1):
        c = ws2.cell(row=i, column=col)
        c.value = val
        c.font = normal_font
        c.border = border
        c.alignment = center
        if i % 2 == 0:
            c.fill = PatternFill("solid", fgColor="EBF3FB")

# ── (a) Ideal Gas
ws2.merge_cells("A14:H14")
ws2["A14"].value = "(a) Ideal Gas Law:  v = RT/P"
ws2["A14"].font = Font(bold=True, size=11, color="FFFFFF")
ws2["A14"].fill = header_fill
ws2["A14"].alignment = center

v_ideal = R * T / P
ws2["A15"].value = "v_ideal (cm³/mol)"
ws2["A15"].font = label_font
ws2["A15"].fill = label_fill
ws2["A15"].border = border
ws2["A15"].alignment = left
ws2.merge_cells("A15:C15")
ws2["D15"].value = round(v_ideal, 4)
ws2["D15"].font = Font(bold=True, size=10, color="FF0000")
ws2["D15"].border = border
ws2["D15"].alignment = center

ws2["A16"].value = "Formula: v = R × T / P = 82.06 × 550 / 270"
ws2["A16"].font = Font(italic=True, size=9)
ws2["A16"].alignment = left
ws2.merge_cells("A16:H16")

# ── (b) Redlich-Kwong EOS
ws2.merge_cells("A18:H18")
ws2["A18"].value = "(b) Redlich-Kwong (RK) Equation of State"
ws2["A18"].font = Font(bold=True, size=11, color="FFFFFF")
ws2["A18"].fill = header_fill
ws2["A18"].alignment = center

rk_headers = ["Component", "yi", "ai (RK)", "bi (RK)", "yi*ai", "yi*bi"]
for col, h in enumerate(rk_headers, 1):
    cell = ws2.cell(row=19, column=col)
    style_header(cell, h, fill=subheader_fill, font=subheader_font)

a_rk_list, b_rk_list, yi_list = [], [], []
for i, (yi, Tc, Pc, omega, Tr) in enumerate(comp_data, start=20):
    a_i = 0.42748 * R**2 * Tc**2.5 / Pc
    b_i = 0.08664 * R * Tc / Pc
    a_rk_list.append(a_i)
    b_rk_list.append(b_i)
    yi_list.append(yi)
    row_vals = [components[i-20][0], yi, round(a_i,2), round(b_i,4), round(yi*a_i,2), round(yi*b_i,4)]
    for col, val in enumerate(row_vals, 1):
        c = ws2.cell(row=i, column=col)
        c.value = val
        c.font = normal_font
        c.border = border
        c.alignment = center
        if i % 2 == 0:
            c.fill = PatternFill("solid", fgColor="EBF3FB")

# Mixing rules RK
a_mix_rk = 0
for i in range(3):
    for j in range(3):
        a_mix_rk += yi_list[i] * yi_list[j] * math.sqrt(a_rk_list[i] * a_rk_list[j])
b_mix_rk = sum(yi_list[i] * b_rk_list[i] for i in range(3))

ws2["A23"].value = "a_mix (RK)"
ws2["A23"].font = label_font; ws2["A23"].fill = label_fill
ws2["A23"].border = border; ws2["A23"].alignment = left
ws2.merge_cells("A23:C23")
ws2["D23"].value = round(a_mix_rk, 2)
ws2["D23"].font = normal_font; ws2["D23"].border = border; ws2["D23"].alignment = center

ws2["A24"].value = "b_mix (RK)"
ws2["A24"].font = label_font; ws2["A24"].fill = label_fill
ws2["A24"].border = border; ws2["A24"].alignment = left
ws2.merge_cells("A24:C24")
ws2["D24"].value = round(b_mix_rk, 4)
ws2["D24"].font = normal_font; ws2["D24"].border = border; ws2["D24"].alignment = center

# Solve RK cubic numerically
def rk_pressure(v, T, a, b, R):
    return R*T/(v-b) - a/(math.sqrt(T)*v*(v+b))

def solve_rk(T, P, a, b, R, v0):
    v = v0
    for _ in range(10000):
        f = rk_pressure(v, T, a, b, R) - P
        dv = 1e-6
        fp = (rk_pressure(v+dv, T, a, b, R) - rk_pressure(v-dv, T, a, b, R)) / (2*dv)
        if abs(fp) < 1e-30:
            break
        v_new = v - f/fp
        if v_new <= b:
            v_new = v * 0.99
        if abs(v_new - v) < 1e-10:
            break
        v = v_new
    return v

v_rk = solve_rk(T, P, a_mix_rk, b_mix_rk, R, v_ideal)

ws2["A25"].value = "v_RK (cm³/mol)  [Goal Seek result]"
ws2["A25"].font = label_font; ws2["A25"].fill = label_fill
ws2["A25"].border = border; ws2["A25"].alignment = left
ws2.merge_cells("A25:C25")
ws2["D25"].value = round(v_rk, 4)
ws2["D25"].font = Font(bold=True, size=10, color="FF0000")
ws2["D25"].border = border; ws2["D25"].alignment = center

# ── (c) RKS EOS
ws2.merge_cells("A27:H27")
ws2["A27"].value = "(c) Redlich-Kwong-Soave (RKS) Equation of State"
ws2["A27"].font = Font(bold=True, size=11, color="FFFFFF")
ws2["A27"].fill = header_fill
ws2["A27"].alignment = center

rks_headers = ["Component", "yi", "m_i", "alpha_i", "ai (RKS)", "bi (RKS)"]
for col, h in enumerate(rks_headers, 1):
    cell = ws2.cell(row=28, column=col)
    style_header(cell, h, fill=subheader_fill, font=subheader_font)

a_rks_list, b_rks_list = [], []
for i, (yi, Tc, Pc, omega, Tr) in enumerate(comp_data, start=29):
    m_i = 0.480 + 1.574*omega - 0.176*omega**2
    alpha_i = (1 + m_i*(1 - math.sqrt(Tr)))**2
    a_i_rks = 0.42748 * R**2 * Tc**2 / Pc * alpha_i
    b_i_rks = 0.08664 * R * Tc / Pc
    a_rks_list.append(a_i_rks)
    b_rks_list.append(b_i_rks)
    row_vals = [components[i-29][0], yi, round(m_i,4), round(alpha_i,4), round(a_i_rks,2), round(b_i_rks,4)]
    for col, val in enumerate(row_vals, 1):
        c = ws2.cell(row=i, column=col)
        c.value = val
        c.font = normal_font
        c.border = border
        c.alignment = center
        if i % 2 == 0:
            c.fill = PatternFill("solid", fgColor="EBF3FB")

a_mix_rks = 0
for i in range(3):
    for j in range(3):
        a_mix_rks += yi_list[i] * yi_list[j] * math.sqrt(a_rks_list[i] * a_rks_list[j])
b_mix_rks = sum(yi_list[i] * b_rks_list[i] for i in range(3))

ws2["A32"].value = "a_mix (RKS)"
ws2["A32"].font = label_font; ws2["A32"].fill = label_fill
ws2["A32"].border = border; ws2["A32"].alignment = left
ws2.merge_cells("A32:C32")
ws2["D32"].value = round(a_mix_rks, 2)
ws2["D32"].font = normal_font; ws2["D32"].border = border; ws2["D32"].alignment = center

ws2["A33"].value = "b_mix (RKS)"
ws2["A33"].font = label_font; ws2["A33"].fill = label_fill
ws2["A33"].border = border; ws2["A33"].alignment = left
ws2.merge_cells("A33:C33")
ws2["D33"].value = round(b_mix_rks, 4)
ws2["D33"].font = normal_font; ws2["D33"].border = border; ws2["D33"].alignment = center


def rks_pressure(v, T, a, b, R):
    return R*T/(v-b) - a/(v*(v+b))

def solve_rks(T, P, a, b, R, v0):
    v = v0
    for _ in range(10000):
        f = rks_pressure(v, T, a, b, R) - P
        dv = 1e-6
        fp = (rks_pressure(v+dv, T, a, b, R) - rks_pressure(v-dv, T, a, b, R)) / (2*dv)
        if abs(fp) < 1e-30:
            break
        v_new = v - f/fp
        if v_new <= b:
            v_new = v * 0.99
        if abs(v_new - v) < 1e-10:
            break
        v = v_new
    return v

v_rks = solve_rks(T, P, a_mix_rks, b_mix_rks, R, v_ideal)

ws2["A34"].value = "v_RKS (cm³/mol)  [Goal Seek result]"
ws2["A34"].font = label_font; ws2["A34"].fill = label_fill
ws2["A34"].border = border; ws2["A34"].alignment = left
ws2.merge_cells("A34:C34")
ws2["D34"].value = round(v_rks, 4)
ws2["D34"].font = Font(bold=True, size=10, color="FF0000")
ws2["D34"].border = border; ws2["D34"].alignment = center

# ── Comparison Table
ws2.merge_cells("A36:H36")
ws2["A36"].value = "Comparison of Results"
ws2["A36"].font = Font(bold=True, size=11, color="FFFFFF")
ws2["A36"].fill = header_fill
ws2["A36"].alignment = center

comp_table_headers = ["Method", "v (cm³/mol)", "Comment"]
for col, h in enumerate(comp_table_headers, 1):
    cell = ws2.cell(row=37, column=col)
    style_header(cell, h, fill=subheader_fill, font=subheader_font)
ws2.merge_cells("C37:H37")

comp_results = [
    ("Ideal Gas Law", round(v_ideal,4), "Assumes no intermolecular forces — overestimates v at high P"),
    ("Redlich-Kwong EOS", round(v_rk,4), "Accounts for attraction & repulsion — more accurate"),
    ("RK-Soave (RKS) EOS", round(v_rks,4), "Includes acentric factor ω → best for polar molecules like NH3"),
]
for i, (method, v_val, comment) in enumerate(comp_results, start=38):
    ws2.cell(row=i, column=1).value = method
    ws2.cell(row=i, column=1).font = label_font
    ws2.cell(row=i, column=1).fill = label_fill
    ws2.cell(row=i, column=1).border = border
    ws2.cell(row=i, column=1).alignment = center
    ws2.cell(row=i, column=2).value = v_val
    ws2.cell(row=i, column=2).font = Font(bold=True, size=10)
    ws2.cell(row=i, column=2).border = border
    ws2.cell(row=i, column=2).alignment = center
    ws2.cell(row=i, column=3).value = comment
    ws2.cell(row=i, column=3).font = Font(italic=True, size=9)
    ws2.cell(row=i, column=3).border = border
    ws2.cell(row=i, column=3).alignment = left
    ws2.merge_cells(f"C{i}:H{i}")
    if i % 2 == 0:
        for col in [1,2]:
            ws2.cell(row=i, column=col).fill = PatternFill("solid", fgColor="EBF3FB")

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 60

# ─────────────────────────────────────────────
# SHEET 3: Q3 - VLE
# ─────────────────────────────────────────────
ws3 = wb.create_sheet("Q3 - VLE")

ws3.merge_cells("A1:H1")
ws3["A1"].value = "Q3 - Benzene-Toluene VLE: Pxy and Txy Diagrams (Raoult's Law)"
ws3["A1"].font = Font(bold=True, size=13, color="FFFFFF")
ws3["A1"].fill = header_fill
ws3["A1"].alignment = center

# Antoine constants
ws3.merge_cells("A3:H3")
ws3["A3"].value = "Antoine Equation: log10(Psat) = A - B/(C+T),  Psat in mmHg, T in °C"
ws3["A3"].font = Font(bold=True, italic=True, size=10)
ws3["A3"].alignment = left

ant_headers = ["Component", "A", "B", "C"]
for col, h in enumerate(ant_headers, 1):
    cell = ws3.cell(row=4, column=col)
    style_header(cell, h, fill=subheader_fill, font=subheader_font)

antoine = [
    ("Benzene",  6.90565, 1211.033, 220.790),
    ("Toluene",  6.95464, 1344.800, 219.482),
]
for i, (name, A, B, C) in enumerate(antoine, start=5):
    row_vals = [name, A, B, C]
    for col, val in enumerate(row_vals, 1):
        c = ws3.cell(row=i, column=col)
        c.value = val
        c.font = normal_font
        c.border = border
        c.alignment = center
        if i % 2 == 0:
            c.fill = PatternFill("solid", fgColor="EBF3FB")

# ── Pxy at T=80°C
ws3.merge_cells("A8:H8")
ws3["A8"].value = "Pxy Diagram Data — Isothermal at T = 80°C"
ws3["A8"].font = Font(bold=True, size=11, color="FFFFFF")
ws3["A8"].fill = header_fill
ws3["A8"].alignment = center

pxy_headers = ["x_B (liq)", "Psat_B (mmHg)", "Psat_T (mmHg)", "P_bubble (mmHg)", "y_B (vap)"]
for col, h in enumerate(pxy_headers, 1):
    cell = ws3.cell(row=9, column=col)
    style_header(cell, h, fill=subheader_fill, font=subheader_font)
    ws3.column_dimensions[get_column_letter(col)].width = 20

T_iso = 80
psat_B_80 = 10**(6.90565 - 1211.033/(220.790 + T_iso))
psat_T_80 = 10**(6.95464 - 1344.800/(219.482 + T_iso))

pxy_data = []
for i, x in enumerate([round(j*0.05,2) for j in range(21)], start=10):
    P_bub = x * psat_B_80 + (1-x) * psat_T_80
    y_B = x * psat_B_80 / P_bub if P_bub > 0 else 0
    pxy_data.append((x, P_bub, y_B))
    row_vals = [x, round(psat_B_80,2), round(psat_T_80,2), round(P_bub,2), round(y_B,4)]
    for col, val in enumerate(row_vals, 1):
        c = ws3.cell(row=i, column=col)
        c.value = val
        c.font = normal_font
        c.border = border
        c.alignment = center
        if i % 2 == 0:
            c.fill = PatternFill("solid", fgColor="EBF3FB")

# Pxy Chart
chart_pxy = LineChart()
chart_pxy.title = "Pxy Diagram — Benzene-Toluene at 80°C"
chart_pxy.style = 10
chart_pxy.y_axis.title = "Pressure (mmHg)"
chart_pxy.x_axis.title = "x_B or y_B (mole fraction Benzene)"
chart_pxy.height = 14
chart_pxy.width = 24

x_ref = Reference(ws3, min_col=1, min_row=10, max_row=30)
bubble_ref = Reference(ws3, min_col=4, min_row=9, max_row=30)
dew_ref = Reference(ws3, min_col=5, min_row=9, max_row=30)

s_bub = openpyxl.chart.Series(bubble_ref, x_ref, title_from_data=True)
s_bub.graphicalProperties.line.solidFill = "0070C0"
s_bub.graphicalProperties.line.width = 20000
s_dew = openpyxl.chart.Series(dew_ref, x_ref, title_from_data=True)
s_dew.graphicalProperties.line.solidFill = "FF0000"
s_dew.graphicalProperties.line.width = 20000
chart_pxy.series.append(s_bub)
chart_pxy.series.append(s_dew)
ws3.add_chart(chart_pxy, "G8")

# ── Txy at P=760 mmHg
ws3.merge_cells("A32:H32")
ws3["A32"].value = "Txy Diagram Data — Isobaric at P = 760 mmHg (1 atm)"
ws3["A32"].font = Font(bold=True, size=11, color="FFFFFF")
ws3["A32"].fill = header_fill
ws3["A32"].alignment = center

txy_headers = ["x_B (liq)", "T_bubble (°C)", "Psat_B (mmHg)", "Psat_T (mmHg)", "y_B (vap)"]
for col, h in enumerate(txy_headers, 1):
    cell = ws3.cell(row=33, column=col)
    style_header(cell, h, fill=subheader_fill, font=subheader_font)

P_total = 760

def bubble_T(xB, P_total, tol=1e-6, max_iter=1000):
    T_lo, T_hi = 60.0, 120.0
    for _ in range(max_iter):
        T_mid = (T_lo + T_hi) / 2
        pB = 10**(6.90565 - 1211.033/(220.790 + T_mid))
        pT = 10**(6.95464 - 1344.800/(219.482 + T_mid))
        P_calc = xB * pB + (1-xB) * pT
        if abs(P_calc - P_total) < tol:
            break
        if P_calc > P_total:
            T_hi = T_mid
        else:
            T_lo = T_mid
    return T_mid

txy_data = []
for i, x in enumerate([round(j*0.05,2) for j in range(21)], start=34):
    T_bub = bubble_T(x, P_total)
    pB = 10**(6.90565 - 1211.033/(220.790 + T_bub))
    pT = 10**(6.95464 - 1344.800/(219.482 + T_bub))
    y_B = x * pB / P_total
    txy_data.append((x, T_bub, y_B))
    row_vals = [round(x,2), round(T_bub,2), round(pB,2), round(pT,2), round(y_B,4)]
    for col, val in enumerate(row_vals, 1):
        c = ws3.cell(row=i, column=col)
        c.value = val
        c.font = normal_font
        c.border = border
        c.alignment = center
        if i % 2 == 0:
            c.fill = PatternFill("solid", fgColor="EBF3FB")

# Txy Chart
chart_txy = LineChart()
chart_txy.title = "Txy Diagram — Benzene-Toluene at 1 atm"
chart_txy.style = 10
chart_txy.y_axis.title = "Temperature (°C)"
chart_txy.x_axis.title = "x_B or y_B (mole fraction Benzene)"
chart_txy.height = 14
chart_txy.width = 24

x_ref2 = Reference(ws3, min_col=1, min_row=34, max_row=54)
tbub_ref = Reference(ws3, min_col=2, min_row=33, max_row=54)
yB_ref = Reference(ws3, min_col=5, min_row=33, max_row=54)

s_tbub = openpyxl.chart.Series(tbub_ref, x_ref2, title_from_data=True)
s_tbub.graphicalProperties.line.solidFill = "0070C0"
s_tbub.graphicalProperties.line.width = 20000
s_tdew = openpyxl.chart.Series(yB_ref, x_ref2, title_from_data=True)
s_tdew.graphicalProperties.line.solidFill = "FF0000"
s_tdew.graphicalProperties.line.width = 20000
chart_txy.series.append(s_tbub)
chart_txy.series.append(s_tdew)
ws3.add_chart(chart_txy, "G32")

# ─────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────
wb.save("CH331_Assignment1.xlsx")
print("✅ Excel file 'CH331_Assignment1.xlsx' created successfully!")
